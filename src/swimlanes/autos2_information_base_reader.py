from domain_model.requirements_guarantees_classes import Mitre_Mitigation, Mitre_Technique, Mitre_Technique_Level_Enum
from openpyxl import load_workbook
import setup


def get_techniques_from_information_base(file_name:str, excel_tab:str) -> list[Mitre_Technique]:
    print("Get all specified MITRE Techniques from AutoS² Information Base in file", file_name, "| Tab:", excel_tab)
    print("Add attacker Skills and Resources to the MITRE Technique")
    all_mitre_techniques:list[Mitre_Technique] = []
    workbook = load_workbook(file_name)
    sheet = workbook[excel_tab]    
    if sheet.cell(row=1, column=1).value != "#" or sheet.cell(row=1, column=2).value != "ICS Technique" or sheet.cell(row=1, column=3).value != "Minimum TAL Skill" or sheet.cell(row=1, column=4).value != "Minimum TAL Resources":
        raise ValueError("Wrong Excel Format?", file_name, excel_tab)
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=1).value is None:
            break
        name = sheet.cell(row=i, column=2).value
        minimum_tal_skill = sheet.cell(row=i, column=3).value
        minimum_tal_resources = sheet.cell(row=i, column=4).value
        all_mitre_techniques.append(Mitre_Technique(name, minimum_tal_skill, minimum_tal_resources))
    if setup.PRINT_RESULTS:
        print("|-- Number of Techniques:   ", len(all_mitre_techniques))
        print("|-- First Technique Details:", all_mitre_techniques[0].name, "| Skill:", all_mitre_techniques[0].minimum_tal_skill, "| Resources:", all_mitre_techniques[0].minimum_tal_resources, "| SL-T:", all_mitre_techniques[0].sl_t, "| Level:", all_mitre_techniques[0].technique_level)
        print("|-- Last Technique Details: ", all_mitre_techniques[-1].name,"| Skill:", all_mitre_techniques[-1].minimum_tal_skill,"| Resources:", all_mitre_techniques[-1].minimum_tal_resources,"| SL-T:", all_mitre_techniques[-1].sl_t, "| Level:", all_mitre_techniques[0].technique_level)
        print("|-- Techniques:             ", [techn.name for techn in all_mitre_techniques])
        print()
    return all_mitre_techniques


def get_mitigations_from_information_base(file_name:str, excel_tab:str) -> list[Mitre_Mitigation]:
    print("Get all MITRE Mitigations with assigned IEC 62443 CR/SR from AutoS² Information Base in file", file_name, "| Tab:", excel_tab)
    print("Get CR/SR-ID per MITRE Technique")
    all_mitre_mitigations:list[Mitre_Mitigation] = []
    workbook = load_workbook(file_name)
    sheet = workbook[excel_tab]
    if sheet.cell(row=1, column=1).value != "ID" or sheet.cell(row=1, column=2).value != "Mitigation" or sheet.cell(row=1, column=3).value != "CR/SR":
        raise ValueError("Wrong Excel Format?", file_name, excel_tab)
    for i in range(2, sheet.max_row+1):
        name = sheet.cell(row=i, column=2).value
        id = sheet.cell(row=i, column=1).value
        cr_sr = sheet.cell(row=i, column=3).value
        all_mitre_mitigations.append(Mitre_Mitigation(name, id, cr_sr))
    for mitigation in all_mitre_mitigations:
        if not mitigation.cr_sr:
            setup.error_list.append("No CR/SR assigned for Mitigation '" + mitigation.name + "'. Not considered in further assessment.")
            if setup.PRINT_RESULTS:
                print(" ! ", setup.error_list[-1])
    # Remove all Mitigations with unknown CR/SR
    all_mitre_mitigations = [item for item in all_mitre_mitigations if item.cr_sr is not None]
    if setup.PRINT_RESULTS:
        print("|-- Number of Mitigations: ", len(all_mitre_mitigations))
        print("|-- First Mitigation:      ", all_mitre_mitigations[0].name, "| ID:", all_mitre_mitigations[0].id, "| CR/SR:", all_mitre_mitigations[0].cr_sr)
        print("|-- Last Mitigation:       ", all_mitre_mitigations[-1].name,"| ID:", all_mitre_mitigations[-1].id,"| CR/SR:", all_mitre_mitigations[-1].cr_sr)
    print()
    
    addressed_cr_sr:list = []
    for mitigation in all_mitre_mitigations:
        addressed_cr_sr.append(mitigation.cr_sr)
    addressed_cr_sr = list(dict.fromkeys(addressed_cr_sr))
    print("Only the following", len(addressed_cr_sr),"CRs/SRs are addressed by MITRE Mitigations and therefore part of the further assessment:")
    print(addressed_cr_sr)
    
    print()
    return all_mitre_mitigations


def assign_mitigations_to_technique(file_name:str, excel_tab:str, all_mitre_techniques:list[Mitre_Technique], all_mitre_mitigations:list[Mitre_Mitigation]) -> list[Mitre_Technique]:
    print("Get all MITRE Mitigations for MITRE Techniques according to MITRE ICS in file", file_name, "| Tab:", excel_tab)
    number_of_techniques_with_mitigations = 0
    for technique in all_mitre_techniques:
        technique:Mitre_Technique
        mitre_mitigations:list[Mitre_Mitigation] = []
        workbook = load_workbook(file_name)
        sheet = workbook[excel_tab]
        if sheet.cell(row=1, column=2).value != "source name" or sheet.cell(row=1, column=6).value != "target name":
            raise ValueError("Wrong Excel Format?", file_name, excel_tab)
        for i in range(2, sheet.max_row+1):
            if sheet.cell(row=i, column=6).value is None:
                break
            if sheet.cell(row=i, column=6).value == technique.name:
                corresponding_mitigation = sheet.cell(row=i, column=2).value
                for mitigation in all_mitre_mitigations:
                    if mitigation.name == corresponding_mitigation:
                        mitre_mitigations.append(mitigation)
        technique.mitigations = mitre_mitigations
        if technique.mitigations == []:
            setup.error_list.append("No Mitigations found for Technique '" + technique.name + "'. Not considered in further assessment.")
            if setup.PRINT_RESULTS:
                print(" ! ", setup.error_list[-1])
        else:
            number_of_techniques_with_mitigations+=1
    if setup.PRINT_RESULTS:
        print("|-- Number of Techniques with Mitigations: ", number_of_techniques_with_mitigations)
        print("|-- Mitigations for ", all_mitre_techniques[0].name, "=", [(mitigation.name, mitigation.cr_sr) for mitigation in all_mitre_techniques[0].mitigations])
        print("|-- Mitigations for ", all_mitre_techniques[-1].name, "=",[(mitigation.name, mitigation.cr_sr) for mitigation in all_mitre_techniques[-1].mitigations])
        print()
    print()
    return all_mitre_techniques


def get_technique_dict_for_cves_from_information_base(file_name:str, excel_tab:str, all_mitre_techniques:list[Mitre_Technique]) -> dict:
    print("Read Techniques for CVEs from AutoS² Information Base in file", file_name, "| Tab:", excel_tab)
    technique_names_for_cves:dict = {}
    workbook = load_workbook(file_name)
    sheet = workbook[excel_tab]
    if sheet.cell(row=1, column=1).value != "#" or sheet.cell(row=1, column=2).value != "Asset" or sheet.cell(row=1, column=3).value != "CVE ID" or sheet.cell(row=1, column=4).value != "Exploitation (ICS Technique)" or sheet.cell(row=1, column=5).value != "Primary Impact (ICS Technique)" or sheet.cell(row=1, column=6).value != "Secondary Impact (ICS Technique)":
        raise ValueError("Wrong Excel Format?", file_name, excel_tab)
    for i in range(2, sheet.max_row+1):
        cve_id = sheet.cell(row=i, column=3).value
        technique_name_and_levels:list[(str, Mitre_Technique_Level_Enum)] = []
        for j in range(4, 7):
            if j == 4:
                level = Mitre_Technique_Level_Enum.EXPLOITATION
            elif j == 5:
                level = Mitre_Technique_Level_Enum.PRIMARYIMPACT
            elif j == 6:
                level = Mitre_Technique_Level_Enum.SECONDAYIMPACT
            else:
                level = Mitre_Technique_Level_Enum.UNKNOWN
            value = sheet.cell(row=i, column=j).value
            if value != "-":
                technique_name_and_level:tuple = (value, level)
                technique_name_and_levels.append(technique_name_and_level)
        technique_names_for_cves[cve_id] = technique_name_and_levels
    techniques_for_cves:dict = {}
    for cve_id, technique_name_and_levels in technique_names_for_cves.items():
        techniques_for_cve:dict = {}
        for technique_name_and_level in technique_name_and_levels:
            technique_found = False
            for mitre_technique in all_mitre_techniques:
                if mitre_technique.name == technique_name_and_level[0]:
                    techniques_for_cve[technique_name_and_level[1]] = mitre_technique
                    technique_found = True
            if not technique_found:
                setup.error_list.append("Technique '" + technique_name_and_level[0] + "' not found in AutoS² Information Base. Not considered in further assessment.")
                if setup.PRINT_RESULTS:
                    print(" ! ", setup.error_list[-1])
        techniques_for_cves[cve_id] = techniques_for_cve
    if setup.PRINT_RESULTS:
        cve_id_example = "CVE-2020-12518"
        cve_level_example = Mitre_Technique_Level_Enum.SECONDAYIMPACT
        print("|-- Example", cve_level_example.name, "for", cve_id_example, "=", techniques_for_cves[cve_id_example][cve_level_example].name)
    return techniques_for_cves
